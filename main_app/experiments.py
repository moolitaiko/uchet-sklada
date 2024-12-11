import openpyxl
import math
from openpyxl.styles import Alignment, Font

def GenerateCheck(sale):
    sale_data = sale.datetime.strftime('%d`%m`%y')
    file_name = f'Checki/{sale.id} от {sale_data}.xlsx'
    wb = openpyxl.Workbook()
    page = wb.active
    page.cell(row=1, column=1).value = sale.seller.owner.company_name
    page.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    # page.cell(row=1, column=1).alignment = Alignment(wrap_text=True)
    # page.cell(row=1, column=1).height = None

    page.cell(row=3, column=1).value = sale.seller.owner.fact_address
    # page.cell(row=3, column=1).alignment = Alignment(wrap_text = True)
    # page.row_dimensions[3].auto_height = True

    page.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)

    page.cell(row=5, column=3).value = f'Товарный чек №{sale.id} от {sale_data}'
    page.merge_cells(start_row=5, start_column=3, end_row=5, end_column=7)
    page.cell(row=5, column=3).alignment = Alignment(horizontal='center')
    page.cell(row=5, column=3).font = Font(bold=True)


    row_number = 8
    page.cell(row=7, column=1).value = '№'
    page.cell(row=7, column=2).value = 'Артикул'
    page.cell(row=7, column=3).value = 'Наименование'
    page.cell(row=7, column=4).value = 'Количество'
    page.cell(row=7, column=5).value = 'Единицы'
    page.cell(row=7, column=6).value = 'Цена'
    page.cell(row=7, column=7).value = 'Сумма'
    row_number = 8
    number_position = 1
    total = 0
    for position in sale.positions.all():
        page.cell(row=row_number, column=1).value = number_position
        page.cell(row=row_number, column=2).value = position.product.article
        page.cell(row=row_number, column=3).value = position.product.name
        page.cell(row=row_number, column=4).value = position.quantity
        page.cell(row=row_number, column=5).value = position.units_of_measurement.units
        page.cell(row=row_number, column=6).value = position.sale_price
        page.cell(row=row_number, column=7).value = position.quantity * position.sale_price
        number_position += 1
        row_number += 1
        total += position.quantity * position.sale_price

    page.cell(row=row_number + 1, column=5).value = 'Сумма чека'
    page.cell(row=row_number + 1, column=5).alignment = Alignment(horizontal='right')
    page.merge_cells(start_row=row_number + 1, start_column=5, end_row=row_number + 1, end_column=6)
    page.cell(row=row_number + 1, column=5).font = Font(bold=True)
    page.cell(row=row_number + 1, column=7).value = total
    page.cell(row=row_number + 2, column = 6).value = 'Скидка'
    page.cell(row=row_number + 2, column=6).alignment = Alignment(horizontal='right')
    page.cell(row=row_number + 2, column = 7).value = sale.discount
    page.cell(row=row_number + 3, column = 6).value = 'Итого'
    page.cell(row=row_number + 3, column=6).font = Font(bold=True)
    page.cell(row=row_number + 3, column=6).alignment = Alignment(horizontal='right')
    if sale.discount_type == 'общая скидка':
        final = total - sale.discount
        page.cell(row=row_number + 3, column = 7).value = final
    else:
        final = total - (total * sale.discount/100)
        page.cell(row=row_number + 3, column=7).value = final

    page.cell(row=row_number + 5, column = 1).value = f'Всего наименований {number_position -1} на сумму {final}'
    page.merge_cells(start_row=row_number + 5, start_column=1, end_row=row_number + 5, end_column=7)
    page.cell(row=row_number + 5, column=1).font = Font(bold=True)

    page.cell(row=row_number + 7, column=1).value = 'Продавец'
    page.cell(row= row_number + 7, column=6).value = f'{sale.seller.user.last_name} {sale.seller.user.first_name}'
    page.merge_cells(start_row=row_number + 7, start_column=6, end_row=row_number + 7, end_column=7)
    page.cell(row=row_number + 8, column=2).value = 'подпись'
    page.cell(row=row_number + 8, column=2).font = Font(size='7')
    page.merge_cells(start_row=row_number + 8, start_column=2, end_row=row_number + 8, end_column=5)
    page.cell(row=row_number + 8, column=6).font = Font(size='7')
    page.cell(row=row_number + 8, column=2).alignment = Alignment(vertical='top', horizontal='center')
    page.cell(row=row_number + 8, column = 6).value = 'расшифвровка'
    page.merge_cells(start_row=row_number + 8, start_column=6, end_row=row_number + 8, end_column=7)
    page.cell(row=row_number + 8, column=6).font = Font(size='7')
    page.cell(row=row_number + 8, column=6).alignment = Alignment(vertical='top', horizontal='center')



    wb.save(file_name)
    return file_name